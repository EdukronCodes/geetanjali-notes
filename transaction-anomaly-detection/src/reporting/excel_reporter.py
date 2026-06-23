from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelReporter:
    def __init__(self, output_dir: Path) -> None:
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write(self, df: pd.DataFrame, run_month: str, precision: float | None) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.output_dir / f"anomaly_report_{run_month}_{timestamp}.xlsx"

        summary = pd.DataFrame(
            [
                {"Metric": "Run Month", "Value": run_month},
                {"Metric": "Total Transactions", "Value": len(df)},
                {"Metric": "Flagged Anomalies", "Value": int(df["final_flag"].sum())},
                {"Metric": "Validation Precision", "Value": precision if precision else "N/A"},
                {"Metric": "Generated At", "Value": datetime.now().isoformat()},
            ]
        )

        flagged = df[df["final_flag"] == 1].sort_values("risk_score", ascending=False)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name="Summary", index=False)
            flagged.to_excel(writer, sheet_name="Flagged Exceptions", index=False)
            df.to_excel(writer, sheet_name="All Scored Transactions", index=False)

            ws = writer.sheets["Flagged Exceptions"]
            header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                ws.column_dimensions[get_column_letter(col)].width = 18

        return filename
